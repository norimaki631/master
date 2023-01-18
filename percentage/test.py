from re import A
import cv2
import numpy as np

import openpyxl

# PC3の場合
# book = openpyxl.load_workbook("D:\\Misaki Sato\\master\\result\\smile_percentage.xlsx")
book = openpyxl.load_workbook("D:\\Misaki Sato\\master\\result\\smile_percentage.xlsx")
sheet = book["test3"]

name = "hina"

# capture = cv2.VideoCapture("D:\\Misaki Sato\\master\\recording\\hon\\mkv\\%s.mkv" % name)
capture = cv2.VideoCapture("D:\\Misaki Sato\\master\\recording\\yobi\\%s.mp4" % name)
capture.set(3,640)# 320 320 640 720
capture.set(4,480)# 180 240  360 405

video_frame_count = capture.get(cv2.CAP_PROP_FRAME_COUNT) # フレーム数を取得する
video_fps = capture.get(cv2.CAP_PROP_FPS)                 # フレームレートを取得する
video_len_sec = video_frame_count / video_fps         # 長さ（秒）を計算する

face_cascade = cv2.CascadeClassifier('./haarcascade_frontalface_default.xml')
smile_cascade = cv2.CascadeClassifier('./haarcascade_smile.xml')

left_sum = 0
left_average = 0
right_sum = 0
right_sum = 0
frame = 0
left_face_count = 0
right_face_count = 0
left_smile_count = 0
right_smile_count = 0
# left_roi = 0
# right_roi = 0
# left_array = np.array([])
# right_array = np.array([])
left_num = 0
right_num = 0

maxRow = sheet.max_row + 1

while capture.isOpened():
    ret, img = capture.read()

    if ret == True:
        # img = cv2.flip(img,1)#鏡表示にするため．
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        faces = face_cascade.detectMultiScale(gray, 1.1, 5, minSize=(100,100))
        for (x,y,w,h) in faces:
            _, width, _ = img.shape
            # 左の人の顔の計算
            if x < width/2:
                cv2.rectangle(img,(x,y),(x+w,y+h),(255, 0, 0),2) # blue
                #Gray画像から，顔領域を切り出す．
                roi_gray = gray[y:y+h, x:x+w]
                left_face_count += 1

                #サイズを縮小
                roi_gray = cv2.resize(roi_gray,(100,100))
                cv2.imshow("roi_gray1",roi_gray) #確認のためサイズ統一させた画像を表示

                # 輝度を規格化
                lmin = roi_gray.min() #輝度の最小値
                lmax = roi_gray.max() #輝度の最大値
                llist = np.ravel(roi_gray)

                # 輝度を標準化
                lmean = np.mean(llist)
                lstd = np.std(llist)

                for index1, item1 in enumerate(roi_gray):
                    for index2, item2 in enumerate(item1) :
                        roi_gray[index1][index2] = int((item2 - lmean)/lstd * 37 + 128)
                #         left_roi += float((item2 - lmean)/lstd)
                # print(left_roi/10000)
                cv2.imshow("roi_gray1_hyoujunka",roi_gray)

                smiles= smile_cascade.detectMultiScale(roi_gray,scaleFactor= 1.2, minNeighbors=0, minSize=(20, 20))#笑顔識別
                if len(smiles) >0 : # 笑顔領域がなければ以下の処理を飛ばす．#if len(smiles) <=0 : continue でもよい．その場合以下はインデント不要
                    left_smile_count += 1
                    # サイズを考慮した笑顔認識
                    smile_neighbors = len(smiles)
                    # print("smile_neighbors=",smile_neighbors) #確認のため認識した近傍矩形数を出力
                    # left_array = np.append(left_array, smile_neighbors)
                    sheet.cell(maxRow,left_face_count+9).value = smile_neighbors
                    left_sum += smile_neighbors
                    left_average = left_sum / left_smile_count
                    # i += 1
                    for(sx,sy,sw,sh) in smiles:
                        cv2.circle(img,(int(x+(sx+sw/2)*w/100),int(y+(sy+sh/2)*h/100)),int(sw/2*w/100), (255*(1.0-smile_neighbors), 0, 255*smile_neighbors),2)#red
                else:
                    # left_array = np.append(left_array, 0)
                    sheet.cell(maxRow,left_face_count+9).value = 0
                        
            # 右の人の顔の計算
            else:
                cv2.rectangle(img,(x,y),(x+w,y+h),(255, 0, 0),2) # blue
                #Gray画像から，顔領域を切り出す．
                roi_gray = gray[y:y+h, x:x+w] 
                right_face_count += 1

                #サイズを縮小
                roi_gray = cv2.resize(roi_gray,(100,100))
                cv2.imshow("roi_gray2",roi_gray) #確認のためサイズ統一させた画像を表示

                llist = np.ravel(roi_gray)
                lmean = np.mean(llist)
                lstd = np.std(llist)

                for index1, item1 in enumerate(roi_gray):
                    for index2, item2 in enumerate(item1) :
                        roi_gray[index1][index2] = int((item2 - lmean)/lstd * 37 + 128)
                        # right_roi += float((item2 - lmean)/lstd)
                cv2.imshow("roi_gray2_hyoujunka",roi_gray)

                smiles= smile_cascade.detectMultiScale(roi_gray,scaleFactor= 1.2, minNeighbors=0, minSize=(20, 20))#笑顔識別
                if len(smiles) >0 : # 笑顔領域がなければ以下の処理を飛ばす．#if len(smiles) <=0 : continue でもよい．その場合以下はインデント不要
                    right_smile_count += 1
                    # サイズを考慮した笑顔認識
                    smile_neighbors = len(smiles)
                    # print("smile_neighbors=",smile_neighbors) #確認のため認識した近傍矩形数を出力
                    # right_array = np.append(right_array, smile_neighbors)
                    sheet.cell(maxRow+1,right_face_count+9).value = smile_neighbors
                    right_sum += smile_neighbors
                    right_average = right_sum / right_smile_count
                    # j += 1
                    for(sx,sy,sw,sh) in smiles:
                        cv2.circle(img,(int(x+(sx+sw/2)*w/100),int(y+(sy+sh/2)*h/100)),int(sw/2*w/100), (255*(1.0-smile_neighbors), 0, 255*smile_neighbors),2)#red                
                else:
                    # right_array = np.append(right_array, 0)
                    sheet.cell(maxRow+1,right_face_count+9).value = 0

        cv2.imshow('img',img)

        # key Operation
        key = cv2.waitKey(5) 
        if key ==27 or key ==ord('q'): #escまたはeキーで終了
            print("left:" + str(left_average))
            print("right:" + str(right_average))
            maxRow = sheet.max_row + 1
            print(maxRow)
            sheet.cell(maxRow,1).value = name
            sheet.cell(maxRow,2).value = left_average
            sheet.cell(maxRow,3).value = right_average
            book.save("D:\\Misaki Sato\\master\\result\\smile_percentage.xlsx")
            break
        
        frame += 1

    else:
        sheet.cell(maxRow,1).value = name # ファイル名
        sheet.cell(maxRow,2).value = video_len_sec # 動画秒数
        sheet.cell(maxRow,3).value = frame/video_len_sec # fps(計算)
        sheet.cell(maxRow,4).value = video_fps # fps(正式)
        sheet.cell(maxRow,5).value = left_average # 生数字平均(左)
        sheet.cell(maxRow,6).value = left_face_count*100/frame # 顔認識率(左)
        sheet.cell(maxRow,7).value = left_smile_count*100/frame # 笑顔認識率(左)
        sheet.cell(maxRow+1,5).value = right_average # 生数字平均(右)
        sheet.cell(maxRow+1,6).value = right_face_count*100/frame # 顔認識率(右)
        sheet.cell(maxRow+1,7).value = right_smile_count*100/frame  # 笑顔認識率(右)
        # while left_num < left_face_count:
        #     sheet.cell(maxRow,left_num+9).value = left_array[left_num]
        # while right_num < right_face_count:
        #     sheet.cell(maxRow+1,right_num+9).value = left_array[right_num]

        # book.save("C:\\Users\\Misaki Sato\\Desktop\\result\\smile_percentage.xlsx")
        book.save("D:\\Misaki Sato\\master\\result\\smile_percentage.xlsx")
        # print(left_face_count)
        # print(right_face_count)

        break

capture.release()
cv2.destroyAllWindows()

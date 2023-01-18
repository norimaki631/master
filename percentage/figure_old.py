from re import A
import cv2
import numpy as np

import openpyxl

# PC3の場合
book = openpyxl.load_workbook("D:\\Misaki Sato\\master\\result\\smile_percentage.xlsx")
sheet = book["zhang"]

name = "zhang"

# capture = cv2.VideoCapture("D:\\Misaki Sato\\master\\recording\\hon\\mkv\\%s.mkv" % name)
capture = cv2.VideoCapture("D:\\Misaki Sato\\master\\recording\\yobi\\%s.mp4" % name)
capture.set(3,640)# 320 320 640 720
capture.set(4,480)# 180 240  360 405

face_cascade = cv2.CascadeClassifier('./haarcascade_frontalface_default.xml')
smile_cascade = cv2.CascadeClassifier('./haarcascade_smile.xml')

i = 1
j = 1

while capture.isOpened():
    ret, img = capture.read()

    if ret == True:
        # img = cv2.flip(img,1)#鏡表示にするため．
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        faces = face_cascade.detectMultiScale(gray, 1.1, 5, minSize=(100,100))
        for (x,y,w,h) in faces:
            _, width, _ = img.shape
            # 左の人の顔の計算
            if x < width/2:
                cv2.rectangle(img,(x,y),(x+w,y+h),(255, 0, 0),2) # blue
                #Gray画像から，顔領域を切り出す．
                roi_gray = gray[y:y+h, x:x+w]

                #サイズを縮小
                roi_gray = cv2.resize(roi_gray,(100,100))
                #cv2.imshow("roi_gray",roi_gray) #確認のためサイズ統一させた画像を表示

                # 輝度で規格化
                lmin = roi_gray.min() #輝度の最小値
                lmax = roi_gray.max() #輝度の最大値
                # print("lmax:" + str(lmax))
                # print("lmin:" + str(lmin))
                for index1, item1 in enumerate(roi_gray):
                    for index2, item2 in enumerate(item1) :
                        roi_gray[index1][index2] = int((item2 - lmin)/(lmax-lmin) * item2)
                # cv2.imshow("roi_gray12",roi_gray)  #確認のため輝度を正規化した画像を表示

                smiles= smile_cascade.detectMultiScale(roi_gray,scaleFactor= 1.2, minNeighbors=0, minSize=(20, 20))#笑顔識別
                if len(smiles) >0 : # 笑顔領域がなければ以下の処理を飛ばす．#if len(smiles) <=0 : continue でもよい．その場合以下はインデント不要
                    # サイズを考慮した笑顔認識
                    smile_neighbors = len(smiles)
                    # print("smile_neighbors=",smile_neighbors) #確認のため認識した近傍矩形数を出力
                    LV = 2/150
                    left_intensityZeroOne = smile_neighbors  * LV 
                    if left_intensityZeroOne > 1.0: 
                        left_intensityZeroOne = 1.0
                    # print(intensityZeroOne) #確認のため強度を出力
                    sheet.cell(1,i).value = left_intensityZeroOne
                    # book.save("C:\\Users\\Misaki Sato\\Desktop\\result\\smile_percentage.xlsx")
                    i += 1
                    for(sx,sy,sw,sh) in smiles:
                        cv2.circle(img,(int(x+(sx+sw/2)*w/100),int(y+(sy+sh/2)*h/100)),int(sw/2*w/100), (255*(1.0-left_intensityZeroOne), 0, 255*left_intensityZeroOne),2)#red
                        
            # 右の人の顔の計算
            else:
                cv2.rectangle(img,(x,y),(x+w,y+h),(255, 0, 0),2) # blue
                #Gray画像から，顔領域を切り出す．
                roi_gray = gray[y:y+h, x:x+w] 

                #サイズを縮小
                roi_gray = cv2.resize(roi_gray,(100,100))
                #cv2.imshow("roi_gray",roi_gray) #確認のためサイズ統一させた画像を表示

                # 輝度で規格化
                lmin = roi_gray.min() #輝度の最小値
                lmax = roi_gray.max() #輝度の最大値
                for index1, item1 in enumerate(roi_gray):
                    for index2, item2 in enumerate(item1) :
                        roi_gray[index1][index2] = int((item2 - lmin)/(lmax-lmin) * item2)
                        # print(right_roi)
                # cv2.imshow("roi_gray2",roi_gray)  #確認のため輝度を正規化した画像を表示

                smiles= smile_cascade.detectMultiScale(roi_gray,scaleFactor= 1.2, minNeighbors=0, minSize=(20, 20))#笑顔識別
                if len(smiles) >0 : # 笑顔領域がなければ以下の処理を飛ばす．#if len(smiles) <=0 : continue でもよい．その場合以下はインデント不要
                    # サイズを考慮した笑顔認識
                    smile_neighbors = len(smiles)
                    # print("smile_neighbors=",smile_neighbors) #確認のため認識した近傍矩形数を出力
                    LV = 2/150
                    right_intensityZeroOne = smile_neighbors  * LV 
                    if right_intensityZeroOne > 1.0: 
                        right_intensityZeroOne = 1.0
                    # print(intensityZeroOne) #確認のため強度を出力
                    sheet.cell(2,j).value = right_intensityZeroOne
                    # book.save("C:\\Users\\Misaki Sato\\Desktop\\result\\smile_percentage.xlsx")
                    j += 1
                    for(sx,sy,sw,sh) in smiles:
                        cv2.circle(img,(int(x+(sx+sw/2)*w/100),int(y+(sy+sh/2)*h/100)),int(sw/2*w/100), (255*(1.0-right_intensityZeroOne), 0, 255*right_intensityZeroOne),2)#red                

        cv2.imshow('img',img)

    else:
        book.save("D:\\Misaki Sato\\master\\result\\smile_percentage.xlsx")
        break

capture.release()
cv2.destroyAllWindows()


